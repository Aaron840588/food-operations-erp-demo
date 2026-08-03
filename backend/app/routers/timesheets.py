from collections import defaultdict
from datetime import datetime
import io
import re
from typing import List

from fastapi import APIRouter, Depends, HTTPException, Query, Response, UploadFile, File
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

from .. import auth, models, schemas
from ..database import get_db
from ..services.proof_images import InvalidProofImage, normalize_proof_image

router = APIRouter(prefix="/timesheets", tags=["Timesheets"])


def _parse_date_range(date_from: str, date_to: str) -> tuple[str, str]:
    try:
        parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Dates must use YYYY-MM-DD.") from exc
    if parsed_to < parsed_from:
        raise HTTPException(status_code=400, detail="End date cannot be before start date.")
    if (parsed_to - parsed_from).days > 366:
        raise HTTPException(status_code=400, detail="Labor reports are limited to 366 days.")
    return parsed_from.isoformat(), parsed_to.isoformat()


def _parse_machine_timestamp(values: dict) -> datetime | None:
    normalized = {str(key).strip().lower().replace("_", " "): str(value).strip() for key, value in values.items()}
    timestamp = next((value for key, value in normalized.items() if key in {"datetime", "date time", "timestamp", "time stamp", "punch time"}), "")
    if not timestamp:
        date_value = next((value for key, value in normalized.items() if key in {"date", "attendance date", "punch date"}), "")
        time_value = next((value for key, value in normalized.items() if key in {"time", "attendance time", "punch time"}), "")
        timestamp = f"{date_value} {time_value}".strip()
    for pattern in (None, "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            return datetime.fromisoformat(timestamp) if pattern is None else datetime.strptime(timestamp, pattern)
        except ValueError:
            continue
    slash_timestamp = re.fullmatch(
        r"(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})(?::(\d{2}))?\s*(AM|PM)?",
        timestamp,
        flags=re.IGNORECASE,
    )
    if slash_timestamp:
        first, second, year, hour, minute, second_value, meridiem = slash_timestamp.groups()
        first_number, second_number = int(first), int(second)
        if first_number <= 12 and second_number <= 12:
            return None
        month, day = (second_number, first_number) if first_number > 12 else (first_number, second_number)
        hour_number = int(hour)
        if meridiem:
            if not 1 <= hour_number <= 12:
                return None
            hour_number = hour_number % 12 + (12 if meridiem.upper() == "PM" else 0)
        try:
            return datetime(int(year), month, day, hour_number, int(minute), int(second_value or 0))
        except ValueError:
            return None
    return None


def _machine_identity(values: dict) -> tuple[str, str]:
    normalized = {str(key).strip().lower().replace("_", " "): str(value).strip() for key, value in values.items()}
    machine_id = next((value for key, value in normalized.items() if key in {"id", "user id", "employee id", "enroll id", "pin", "no."}), "")
    name = next((value for key, value in normalized.items() if key in {"name", "employee name", "user name", "username"}), "")
    return machine_id, name or machine_id


@router.get("", response_model=schemas.TimesheetPage)
def get_timesheets(
    limit: int = Query(default=50, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    query = db.query(models.TimesheetEntry)
    if current_user.role != "owner":
        query = query.filter(models.TimesheetEntry.employee_user_id == current_user.id)
    total = query.count()
    items = query.order_by(
        models.TimesheetEntry.work_date.desc(), models.TimesheetEntry.clock_in.desc()
    ).offset(offset).limit(limit).all()
    return schemas.TimesheetPage(items=items, total=total, limit=limit, offset=offset)


@router.get("/labor-summary", response_model=schemas.TimesheetLaborSummary, dependencies=[Depends(auth.require_owner)])
def get_labor_summary(
    date_from: str = Query(pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: str = Query(pattern=r"^\d{4}-\d{2}-\d{2}$"),
    db: Session = Depends(get_db),
):
    start_date, end_date = _parse_date_range(date_from, date_to)
    entries = db.query(models.TimesheetEntry).options(
        joinedload(models.TimesheetEntry.employee)
    ).filter(
        models.TimesheetEntry.review_status == "Approved",
        models.TimesheetEntry.work_date >= start_date,
        models.TimesheetEntry.work_date <= end_date,
    ).order_by(models.TimesheetEntry.employee_name.asc()).all()

    employee_rows: dict[str, dict] = {}
    for entry in entries:
        key = str(entry.employee_user_id) if entry.employee_user_id is not None else f"machine:{entry.machine_employee_id or entry.employee_name.lower()}"
        row = employee_rows.setdefault(key, {
            "employee_user_id": entry.employee_user_id,
            "employee_name": entry.employee_name,
            "hourly_rate": entry.hourly_rate,
            "approved_hours": 0.0,
            "labor_cost": 0.0,
            "allocated_hours": 0.0,
            "unallocated_hours": 0.0,
            "missing_rate_hours": 0.0,
        })
        hours = entry.duration_hours
        row["approved_hours"] += hours
        row["labor_cost"] += entry.labor_cost
        if hours > 0 and entry.hourly_rate <= 0:
            row["missing_rate_hours"] += hours
        elif hours > 0 and entry.production_plan_id:
            row["allocated_hours"] += hours
        elif hours > 0:
            row["unallocated_hours"] += hours

    employees = []
    for row in employee_rows.values():
        employees.append({
            **row,
            "approved_hours": round(row["approved_hours"], 2),
            "labor_cost": round(row["labor_cost"], 2),
            "allocated_hours": round(row["allocated_hours"], 2),
            "unallocated_hours": round(row["unallocated_hours"], 2),
            "missing_rate_hours": round(row["missing_rate_hours"], 2),
        })

    return {
        "date_from": start_date,
        "date_to": end_date,
        "approved_hours": round(sum(row["approved_hours"] for row in employees), 2),
        "total_labor_cost": round(sum(row["labor_cost"] for row in employees), 2),
        "allocated_hours": round(sum(row["allocated_hours"] for row in employees), 2),
        "unallocated_hours": round(sum(row["unallocated_hours"] for row in employees), 2),
        "missing_rate_hours": round(sum(row["missing_rate_hours"] for row in employees), 2),
        "employees": employees,
    }


@router.post("/manual", response_model=schemas.TimesheetEntryOut)
def create_manual_timesheet(payload: schemas.TimesheetManualCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    if payload.clock_in.date().isoformat() != payload.work_date:
        raise HTTPException(status_code=400, detail="Work date must match the clock-in date.")
    if payload.clock_out and payload.clock_out < payload.clock_in:
        raise HTTPException(status_code=400, detail="Clock-out cannot be before clock-in.")
    existing = db.query(models.TimesheetEntry).filter(
        models.TimesheetEntry.client_reference == payload.client_reference,
        models.TimesheetEntry.employee_user_id == current_user.id,
    ).first()
    if existing:
        return existing
    try:
        proof_image_data, proof_image_type = normalize_proof_image(payload.proof_image_data, payload.proof_image_type)
    except InvalidProofImage as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    entry = models.TimesheetEntry(
        client_reference=payload.client_reference,
        employee_user_id=current_user.id,
        employee_name=payload.employee_name.strip() if current_user.role == "owner" and payload.employee_name else current_user.username,
        work_date=payload.work_date,
        clock_in=payload.clock_in,
        clock_out=payload.clock_out,
        source="manual",
        review_status="Pending",
        proof_image_data=proof_image_data,
        proof_image_type=proof_image_type,
        notes=payload.notes,
        imported_by_user_id=current_user.id,
    )
    db.add(entry)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        existing = db.query(models.TimesheetEntry).filter(
            models.TimesheetEntry.client_reference == payload.client_reference,
            models.TimesheetEntry.employee_user_id == current_user.id,
        ).first()
        if existing:
            return existing
        raise
    db.refresh(entry)
    return entry


@router.get("/{entry_id}/proof", response_model=schemas.TimesheetProofOut)
def get_timesheet_proof(
    entry_id: int,
    response: Response,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    entry = db.query(models.TimesheetEntry).filter(models.TimesheetEntry.id == entry_id).first()
    if not entry or (current_user.role != "owner" and entry.employee_user_id != current_user.id):
        raise HTTPException(status_code=404, detail="Timesheet proof not found")
    if not entry.proof_image_data or not entry.proof_image_type:
        raise HTTPException(status_code=404, detail="Timesheet proof not found")
    response.headers["Cache-Control"] = "private, no-store"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return schemas.TimesheetProofOut(data_url=entry.proof_image_data, mime_type=entry.proof_image_type)


@router.post("/import", response_model=List[schemas.TimesheetEntryOut], dependencies=[Depends(auth.require_owner)])
def import_machine_timesheets(payload: schemas.TimesheetImportCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.require_owner)):
    grouped: dict[tuple[str, str], list[datetime]] = defaultdict(list)
    names: dict[tuple[str, str], str] = {}
    invalid_rows: list[int] = []
    for row_number, row in enumerate(payload.rows, start=2):
        timestamp = _parse_machine_timestamp(row.values)
        machine_id, name = _machine_identity(row.values)
        if timestamp and (machine_id or name):
            key = (machine_id or name, timestamp.date().isoformat())
            grouped[key].append(timestamp)
            names[key] = name
        else:
            invalid_rows.append(row_number)
    if invalid_rows:
        examples = ", ".join(str(row) for row in invalid_rows[:5])
        suffix = "…" if len(invalid_rows) > 5 else ""
        raise HTTPException(
            status_code=400,
            detail=(
                f"Import stopped before saving because row(s) {examples}{suffix} have a missing identity, "
                "invalid timestamp, or an ambiguous numeric date. Use YYYY-MM-DD or an unambiguous date."
            ),
        )
    if not grouped:
        raise HTTPException(status_code=400, detail="No valid Deli attendance rows found. Export a CSV with ID/Name and Date + Time columns.")
    created = []
    users = {user.username.lower(): user for user in db.query(models.User).all()}
    for (machine_id, work_date), punches in grouped.items():
        punches.sort()
        employee = users.get(names[(machine_id, work_date)].lower()) or users.get(machine_id.lower())
        existing = db.query(models.TimesheetEntry).filter(
            models.TimesheetEntry.source == "machine",
            models.TimesheetEntry.machine_employee_id == machine_id,
            models.TimesheetEntry.work_date == work_date,
        ).first()
        if existing:
            existing.clock_in, existing.clock_out = punches[0], punches[-1] if len(punches) > 1 else None
            if employee and existing.employee_user_id is None:
                existing.employee_user_id = employee.id
            if employee and employee.hourly_rate > 0 and (existing.approved_hourly_rate is None or existing.approved_hourly_rate <= 0):
                existing.approved_hourly_rate = employee.hourly_rate
            created.append(existing)
            continue
        entry = models.TimesheetEntry(
            employee_user_id=employee.id if employee else None,
            employee_name=names[(machine_id, work_date)],
            machine_employee_id=machine_id,
            work_date=work_date,
            clock_in=punches[0],
            clock_out=punches[-1] if len(punches) > 1 else None,
            source="machine",
            review_status="Approved",
            approved_hourly_rate=employee.hourly_rate if employee and employee.hourly_rate > 0 else None,
            imported_by_user_id=current_user.id,
        )
        db.add(entry)
        created.append(entry)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        resolved = []
        for machine_id, work_date in grouped:
            existing = db.query(models.TimesheetEntry).filter(
                models.TimesheetEntry.source == "machine",
                models.TimesheetEntry.machine_employee_id == machine_id,
                models.TimesheetEntry.work_date == work_date,
            ).first()
            if existing:
                resolved.append(existing)
        if len(resolved) == len(grouped):
            return resolved
        raise
    for entry in created:
        db.refresh(entry)
    return created


@router.patch("/{entry_id}/review", response_model=schemas.TimesheetEntryOut, dependencies=[Depends(auth.require_owner)])
def review_manual_timesheet(entry_id: int, payload: schemas.TimesheetReviewUpdate, db: Session = Depends(get_db)):
    entry = db.query(models.TimesheetEntry).filter(models.TimesheetEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Timesheet entry not found")
    entry.review_status = payload.review_status
    entry.approved_hourly_rate = (
        entry.employee.hourly_rate
        if payload.review_status == "Approved" and entry.employee and entry.employee.hourly_rate > 0
        else None
    )
    db.commit()
    db.refresh(entry)
    return entry


@router.patch("/{entry_id}/allocation", response_model=schemas.TimesheetEntryOut, dependencies=[Depends(auth.require_owner)])
def allocate_timesheet(entry_id: int, payload: schemas.TimesheetAllocationUpdate, db: Session = Depends(get_db)):
    entry = db.query(models.TimesheetEntry).filter(models.TimesheetEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Timesheet entry not found")
    if entry.review_status != "Approved":
        raise HTTPException(status_code=400, detail="Only approved time can be allocated to production.")
    if payload.production_plan_id is not None:
        if (entry.approved_hourly_rate is None or entry.approved_hourly_rate <= 0) and entry.employee and entry.employee.hourly_rate > 0:
            entry.approved_hourly_rate = entry.employee.hourly_rate
        if entry.hourly_rate <= 0:
            raise HTTPException(status_code=400, detail="Set an employee hourly rate before allocating labor.")
        plan = db.query(models.ProductionPlan).filter(models.ProductionPlan.id == payload.production_plan_id).first()
        if not plan:
            raise HTTPException(status_code=404, detail="Production plan not found")
    entry.production_plan_id = payload.production_plan_id
    db.commit()
    db.refresh(entry)
    return entry


# ----------------------------------------------------
# TIMESHEET EXCEL CALCULATOR PARSER & ROUTES
# ----------------------------------------------------
def parse_timesheet_excel_bytes(file_bytes: bytes) -> dict:
    import openpyxl
    from datetime import datetime, time
    
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}

    def format_time(t) -> str:
        if isinstance(t, time):
            return t.strftime("%H:%M")
        elif isinstance(t, datetime):
            return t.strftime("%H:%M")
        elif isinstance(t, str):
            return t
        return ""

    def format_date(d) -> str:
        if isinstance(d, datetime):
            return d.strftime("%Y-%m-%d")
        elif isinstance(d, str):
            return d.split()[0] if " " in d else d
        return ""

    for sheetname in wb.sheetnames:
        if sheetname not in ('Che', 'Karen', 'Ren', 'Rica'):
            continue
        ws = wb[sheetname]
        
        staff_name = sheetname
        staff_title = ws['B1'].value or f"Staff: {sheetname}"
        if isinstance(staff_title, str) and "Staff:" in staff_title:
            staff_name = staff_title.replace("Staff:", "").strip()
            
        periods = []
        
        # Scan to find "Payroll Period" positions
        left_period_cell = None
        right_period_cell = None
        
        for r in range(1, 20):
            for c in range(1, 25):
                val = ws.cell(row=r, column=c).value
                if val == "Payroll Period":
                    period_name = ws.cell(row=r, column=c+1).value
                    if c < 8:
                        left_period_cell = (r, c, period_name)
                    else:
                        right_period_cell = (r, c, period_name)
                        
        # Parse left period
        if left_period_cell:
            r_start, c_start, period_name = left_period_cell
            rate = ws.cell(row=8, column=c_start+2).value # Col D
            hours_per_shift = ws.cell(row=8, column=c_start+4).value # Col F
            standard_working_hours = ws.cell(row=10, column=c_start+1).value # Col C
            hourly_rate = ws.cell(row=10, column=c_start+4).value # Col F (Rate/hour)
            
            shifts = []
            for r in range(r_start + 2, 43):
                date_val = ws.cell(row=r, column=c_start).value
                if date_val is None:
                    continue
                if not (isinstance(date_val, datetime) or (isinstance(date_val, str) and len(date_val.strip()) > 0)):
                    continue
                
                date_str = format_date(date_val)
                if isinstance(date_val, str) and not any(char.isdigit() for char in date_val):
                    continue
                    
                start_val = ws.cell(row=r, column=c_start+1).value
                end_val = ws.cell(row=r, column=c_start+2).value
                total_hours = ws.cell(row=r, column=c_start+3).value
                working_days = ws.cell(row=r, column=c_start+4).value
                total_pay = ws.cell(row=r, column=c_start+5).value
                
                shifts.append({
                    "date": date_str,
                    "start": format_time(start_val) if start_val else "",
                    "end": format_time(end_val) if end_val else "",
                    "total_hours": float(total_hours) if total_hours is not None else None,
                    "working_days": float(working_days) if working_days is not None else None,
                    "total_pay": float(total_pay) if total_pay is not None else None
                })
                
            summary = {
                "total_hours": None,
                "working_days": None,
                "paid_work": None,
                "allowances": [],
                "total_pay": None,
                "status": None,
                "remarks": None
            }
            for r in range(25, 53):
                label = ws.cell(row=r, column=c_start).value
                label_b = ws.cell(row=r, column=c_start+1).value
                val = ws.cell(row=r, column=c_start+3).value
                
                if label_b == "Total Payroll Period Hours" or label == "Total Payroll Period Hours":
                    summary["total_hours"] = float(val) if val is not None else None
                elif label_b == "Total Payroll Working Days" or label == "Total Payroll Working Days":
                    summary["working_days"] = float(val) if val is not None else None
                elif label_b == "Paid Work" or label == "Paid Work":
                    summary["paid_work"] = float(val) if val is not None else None
                elif label_b == "Transpo Allowance" or label == "Transpo Allowance":
                    summary["allowances"].append({"label": "Transpo Allowance", "amount": float(val) if val is not None else None})
                elif label_b == "Service+ Reimbursement" or label == "Service+ Reimbursement":
                    summary["allowances"].append({"label": "Service+ Reimbursement", "amount": float(val) if val is not None else None})
                    next_val = ws.cell(row=r+1, column=c_start+3).value
                    if next_val and isinstance(next_val, (int, float)):
                        summary["allowances"].append({"label": "Reimbursement 2", "amount": float(next_val)})
                elif label_b in ("Total Pay ", "Total Pay", "Subtotal Pay ") or label in ("Total Pay ", "Total Pay", "Subtotal Pay "):
                    summary["total_pay"] = float(val) if val is not None else None
                elif label_b == "Status" or label == "Status":
                    summary["status"] = ws.cell(row=r, column=c_start+3).value
                elif label_b == "Remarks" or label == "Remarks":
                    summary["remarks"] = ws.cell(row=r, column=c_start+3).value
                    
            periods.append({
                "period_name": str(period_name) if period_name else "Period",
                "side": "left",
                "rate": float(rate) if rate is not None else None,
                "hours_per_shift": float(hours_per_shift) if hours_per_shift is not None else None,
                "standard_working_hours": float(standard_working_hours) if standard_working_hours is not None else None,
                "hourly_rate": float(hourly_rate) if hourly_rate is not None else None,
                "shifts": shifts,
                "summary": summary
            })
            
        # Parse right period
        if right_period_cell:
            r_start, c_start, period_name = right_period_cell
            rate = ws.cell(row=8, column=c_start+2).value
            hours_per_shift = ws.cell(row=8, column=c_start+4).value
            standard_working_hours = ws.cell(row=10, column=c_start+1).value
            hourly_rate = ws.cell(row=10, column=c_start+4).value
            
            shifts = []
            for r in range(r_start + 2, 43):
                date_val = ws.cell(row=r, column=c_start).value
                if date_val is None:
                    continue
                if not (isinstance(date_val, datetime) or (isinstance(date_val, str) and len(date_val.strip()) > 0)):
                    continue
                date_str = format_date(date_val)
                if isinstance(date_val, str) and not any(char.isdigit() for char in date_val):
                    continue
                    
                start_val = ws.cell(row=r, column=c_start+1).value
                end_val = ws.cell(row=r, column=c_start+2).value
                total_hours = ws.cell(row=r, column=c_start+3).value
                working_days = ws.cell(row=r, column=c_start+4).value
                total_pay = ws.cell(row=r, column=c_start+5).value
                
                shifts.append({
                    "date": date_str,
                    "start": format_time(start_val) if start_val else "",
                    "end": format_time(end_val) if end_val else "",
                    "total_hours": float(total_hours) if total_hours is not None else None,
                    "working_days": float(working_days) if working_days is not None else None,
                    "total_pay": float(total_pay) if total_pay is not None else None
                })
                
            summary = {
                "total_hours": None,
                "working_days": None,
                "paid_work": None,
                "allowances": [],
                "total_pay": None,
                "status": None,
                "remarks": None
            }
            for r in range(25, 53):
                label = ws.cell(row=r, column=c_start).value
                label_alt = ws.cell(row=r, column=c_start+1).value
                val = ws.cell(row=r, column=c_start+3).value
                
                is_hours = label == "Total Payroll Period Hours" or label_alt == "Total Payroll Period Hours"
                is_days = label == "Total Payroll Working Days" or label_alt == "Total Payroll Working Days"
                is_work = label == "Paid Work" or label_alt == "Paid Work"
                is_incentive = "Incentive" in str(label or '') or "Incentive" in str(label_alt or '')
                is_pay = label in ("Total Pay ", "Total Pay", "Subtotal Pay ") or label_alt in ("Total Pay ", "Total Pay", "Subtotal Pay ")
                is_paid = label == "Paid" or label_alt == "Paid"
                is_status = label == "Status" or label_alt == "Status"
                is_remarks = label == "Remarks" or label_alt == "Remarks"
                
                if is_hours:
                    summary["total_hours"] = float(val) if val is not None else None
                elif is_days:
                    summary["working_days"] = float(val) if val is not None else None
                elif is_work:
                    summary["paid_work"] = float(val) if val is not None else None
                elif is_incentive:
                    summary["allowances"].append({"label": str(label or label_alt), "amount": float(val) if val is not None else None})
                elif is_paid:
                    summary["allowances"].append({"label": "Paid (Deduction)", "amount": float(val) if val is not None else None})
                elif is_pay:
                    summary["total_pay"] = float(val) if val is not None else None
                elif is_status:
                    summary["status"] = ws.cell(row=r, column=c_start+3).value or ws.cell(row=r, column=c_start+5).value
                elif is_remarks:
                    summary["remarks"] = ws.cell(row=r, column=c_start+3).value or ws.cell(row=r, column=c_start+5).value

            periods.append({
                "period_name": str(period_name) if period_name else "Period",
                "side": "right",
                "rate": float(rate) if rate is not None else None,
                "hours_per_shift": float(hours_per_shift) if hours_per_shift is not None else None,
                "standard_working_hours": float(standard_working_hours) if standard_working_hours is not None else None,
                "hourly_rate": float(hourly_rate) if hourly_rate is not None else None,
                "shifts": shifts,
                "summary": summary
            })

        # Parse Cash Advances
        ca_col = None
        ca_row = None
        for r in range(1, 15):
            for c in range(1, 25):
                if ws.cell(row=r, column=c).value == "CASH ADVANCE":
                    ca_col = c
                    ca_row = r
                    break
            if ca_col:
                break
                
        cash_advances = []
        if ca_col and ca_row:
            for r in range(ca_row + 2, 45):
                date_val = ws.cell(row=r, column=ca_col).value
                amount_val = ws.cell(row=r, column=ca_col+1).value
                status_val = ws.cell(row=r, column=ca_col+2).value
                
                if amount_val is not None and isinstance(amount_val, (int, float)):
                    cash_advances.append({
                        "date": format_date(date_val) if date_val else "",
                        "amount": float(amount_val),
                        "status": str(status_val) if status_val is not None else None
                    })
                    
        result[sheetname] = {
            "employee_name": staff_name,
            "periods": periods,
            "cash_advances": cash_advances
        }

    return {"employees": result}


@router.get("/calculator", response_model=schemas.TimesheetCalculatorResponse)
def get_calculator_data():
    import os
    paths = [
        r"C:\Users\aaron\Downloads\Staff Timesheet Calculator.xlsx",
        "backend/app/resources/Staff Timesheet Calculator.xlsx",
        "app/resources/Staff Timesheet Calculator.xlsx"
    ]
    filepath = None
    for p in paths:
        if os.path.exists(p):
            filepath = p
            break
    if not filepath:
        raise HTTPException(status_code=404, detail="Timesheet calculator Excel file not found.")
    try:
        with open(filepath, "rb") as f:
            content = f.read()
        return parse_timesheet_excel_bytes(content)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to parse timesheet calculator: {str(exc)}")


@router.post("/calculator/upload", response_model=schemas.TimesheetCalculatorResponse)
async def upload_calculator_file(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx Excel files are supported.")
    try:
        content = await file.read()
        return parse_timesheet_excel_bytes(content)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to parse uploaded timesheet calculator: {str(exc)}")
