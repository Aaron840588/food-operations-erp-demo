import unittest
from unittest.mock import patch, MagicMock, mock_open
import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app import auth, main, models
from app.database import Base, get_db

def make_mock_sheet(name):
    cells = {
        (1, 2): f"Staff: {name}", # ws['B1'] -> row=1, col=2
        (12, 1): "Payroll Period",
        (12, 2): "2026-07-01 TO 2026-07-15",
        (8, 3): 500.0,
        (8, 5): 8.0,
        (10, 2): 8.0,
        (10, 5): 62.5,
        
        (12, 8): "Payroll Period",
        (12, 9): "2026-07-16 TO 2026-07-31",
        (8, 10): 500.0,
        (8, 12): 8.0,
        (10, 9): 8.0,
        (10, 12): 62.5,
        
        (2, 14): "CASH ADVANCE",
    }
    # Shifts start at r_start + 2 = 14
    for r in range(14, 29):
        day = r - 13
        cells[(r, 1)] = f"2026-07-{day:02d}"
        cells[(r, 2)] = "08:00"
        cells[(r, 3)] = "17:00"
        cells[(r, 4)] = 8.0
        cells[(r, 5)] = 1.0
        cells[(r, 6)] = 500.0
        
        cells[(r, 8)] = f"2026-07-{(day+15):02d}"
        cells[(r, 9)] = "08:00"
        cells[(r, 10)] = "17:00"
        cells[(r, 11)] = 8.0
        cells[(r, 12)] = 1.0
        cells[(r, 13)] = 500.0

    # Summary Left
    cells[(30, 1)] = "Total Payroll Period Hours"
    cells[(30, 4)] = 120.0
    cells[(31, 1)] = "Total Payroll Working Days"
    cells[(31, 4)] = 15.0
    cells[(32, 1)] = "Paid Work"
    cells[(32, 4)] = 7500.0
    cells[(33, 1)] = "Transpo Allowance"
    cells[(33, 4)] = 100.0
    cells[(34, 1)] = "Total Pay"
    cells[(34, 4)] = 7600.0
    cells[(35, 1)] = "Status"
    cells[(35, 4)] = "Paid"
    cells[(36, 1)] = "Remarks"
    cells[(36, 4)] = "None"

    # Summary Right
    cells[(30, 8)] = "Total Payroll Period Hours"
    cells[(30, 11)] = 120.0
    cells[(31, 8)] = "Total Payroll Working Days"
    cells[(31, 11)] = 15.0
    cells[(32, 8)] = "Paid Work"
    cells[(32, 11)] = 7500.0
    cells[(33, 8)] = "Total Pay"
    cells[(33, 11)] = 7500.0
    cells[(34, 8)] = "Status"
    cells[(34, 11)] = "Paid"

    # Cash advances (row 4, columns 14, 15, 16)
    cells[(4, 14)] = "2026-07-05"
    cells[(4, 15)] = 1000.0
    cells[(4, 16)] = "Done"

    class MockCell:
        def __init__(self, val):
            self.value = val

    class MockSheet:
        def __getitem__(self, key):
            if key == 'B1':
                return MockCell(cells.get((1, 2)))
            return MockCell(None)

        def cell(self, row, column):
            val = cells.get((row, column))
            return MockCell(val)

    return MockSheet()

class TimesheetCalculatorTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.session_factory = sessionmaker(bind=self.engine)

        def override_db():
            with self.session_factory() as db:
                yield db

        main.app.dependency_overrides[get_db] = override_db
        self.owner = models.User(id=3, username="owner", role="owner", is_active=True)
        main.app.dependency_overrides[auth.get_current_user] = lambda: self.owner
        main.app.dependency_overrides[auth.require_owner] = lambda: self.owner
        self.client = TestClient(main.app)

    def tearDown(self):
        main.app.dependency_overrides.clear()
        self.client.close()
        self.engine.dispose()

    @patch("openpyxl.load_workbook")
    @patch("app.routers.timesheets.open", new_callable=mock_open, read_data=b"fake excel file data")
    @patch("os.path.exists", return_value=True)
    def test_get_calculator_endpoint_success(self, mock_exists, mock_open_file, mock_load_wb):
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["Che", "Cash Advance Summary", "Settings"]
        mock_wb.__getitem__.side_effect = lambda name: make_mock_sheet(name)
        mock_load_wb.return_value = mock_wb

        response = self.client.get("/timesheets/calculator")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        
        self.assertIn("Che", data["employees"])
        che = data["employees"]["Che"]
        self.assertEqual(che["employee_name"], "Che")
        self.assertEqual(len(che["periods"]), 2)
        
        # Verify first period
        p1 = che["periods"][0]
        self.assertEqual(p1["period_name"], "2026-07-01 TO 2026-07-15")
        self.assertEqual(len(p1["shifts"]), 15)
        self.assertEqual(p1["shifts"][0]["date"], "2026-07-01")
        self.assertEqual(p1["summary"]["total_hours"], 120.0)
        self.assertEqual(len(p1["summary"]["allowances"]), 1)
        self.assertEqual(p1["summary"]["allowances"][0]["label"], "Transpo Allowance")
        self.assertEqual(p1["summary"]["allowances"][0]["amount"], 100.0)

        # Verify cash advances
        self.assertEqual(len(che["cash_advances"]), 1)
        self.assertEqual(che["cash_advances"][0]["date"], "2026-07-05")
        self.assertEqual(che["cash_advances"][0]["amount"], 1000.0)
        self.assertEqual(che["cash_advances"][0]["status"], "Done")

    @patch("openpyxl.load_workbook")
    @patch("shutil.copyfileobj")
    @patch("builtins.open")
    def test_upload_calculator_endpoint_success(self, mock_open, mock_copyfileobj, mock_load_wb):
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["Che"]
        mock_wb.__getitem__.side_effect = lambda name: make_mock_sheet(name)
        mock_load_wb.return_value = mock_wb

        # Perform the multipart upload post request
        file_content = b"fake excel file data"
        response = self.client.post(
            "/timesheets/calculator/upload",
            files={"file": ("Staff Timesheet Calculator.xlsx", file_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn("Che", data["employees"])
        self.assertEqual(data["employees"]["Che"]["employee_name"], "Che")
